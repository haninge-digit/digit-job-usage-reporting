import os
import base64
import time
import logging

from datetime import date, timedelta

from elasticsearch import Elasticsearch
from elasticsearch.exceptions import NotFoundError

from jinja2 import Environment, FileSystemLoader

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.cell import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook

from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication

from msgraph.core import GraphClient
from azure.identity import ClientSecretCredential


""" 
Environment
"""
ES_URL = os.getenv('ES_URL',"elasticsearch-master.camunda-zeebe:9200")
REPORT_RECIPIENT  = os.getenv('REPORT_RECIPIENT',"hakan.persson@haninge.se")

TEMPLATE_URL = os.getenv('TEMPLATE_URL',"https://raw.githubusercontent.com/haninge-digit/digit-public-jinja-templates/main")

AD_TENANT_ID = os.getenv('AD_TENANT_ID')
AD_CLIENT_ID = os.getenv('AD_CLIENT_ID')
AD_CLIENT_SECRET = os.getenv('AD_CLIENT_SECRET')

DEBUG_MODE = os.getenv('DEBUG','false') == "true"                       # Global DEBUG logging
LOGFORMAT = "%(asctime)s %(funcName)-10s [%(levelname)s] %(message)s"   # Log format


""" 
Get data from elastic
"""
def get_data(datelist):
    es = Elasticsearch(ES_URL)
    # print(datelist)
    data = {}
    for day in datelist:
        try:
            resp = es.search(index=f"zeebe-record_process-instance-creation_8.1.2_{day}", size=1000, query={"match_all": {}})
            counters = {}
            for hit in resp['hits']['hits']:
                process = hit['_source']['value']['bpmnProcessId']
                if "_worker" not in process:
                    if process not in counters:
                        counters[process] = 1
                    else:
                        counters[process] += 1
            data[day] = counters
        except NotFoundError as e:
            data[day] = {}
    return data


"""
Make a HTML summary
"""
def mk_html(data, header, template):
    process_sums = {}
    for counters in data.values():
        for process,cnt in counters.items():
            if process not in process_sums:
                process_sums[process] = cnt
            else:
                process_sums[process] += cnt
    
    html = template.render(header=header, sums=process_sums)
    return html


"""
Make Excel with details
"""
def mk_excel(data, sheet_title):
    normal_font=Font(name="Arial", size=11, bold=False)
#     bold_font=Font(name="Arial", size=11, bold=True)
    yellow_fill = PatternFill("solid", fgColor="FFFF00")

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    columns = [("Datum",15)]
    plist = []
    for p in data.values():
        for process in p.keys():
            if process not in plist:
                plist.append(process)
                columns.append((process,20))

    for col in range(len(columns)):
        ws.cell(row=1,column=col+1).value = columns[col][0]
        ws.cell(row=1,column=col+1).font = normal_font
        ws.cell(row=1,column=col+1).fill = yellow_fill
        ws.column_dimensions[get_column_letter((col+1))].width = columns[col][1]
    r = 2

    for day, p in data.items():
        ws.cell(row=r,column=1).value = day
        for process, cnt in p.items():
            ws.cell(row=r,column=2+plist.index(process)).value = cnt
        r += 1

    xlsx_data = save_virtual_workbook(wb)
    # with open("lista.xlsx","wb") as f:
    #     f.write(xlsx_data)
    # f.close()
    return xlsx_data


"""
Send report as mail
"""
def send_mail(subject, html, xlsx_data=None):
        message = MIMEMultipart('alternative')          # Create a MIME message
        message["From"] = "NoReply@haninge.se"
        message["To"] = REPORT_RECIPIENT
        message["Subject"] = subject
        message.attach(MIMEText("Kontakta digit@haninge.se om du ser den här texten!", 'plain'))
        message.attach(MIMEText(html, 'html'))      # Add the HTML formatted content
        if xlsx_data:
            attachment = MIMEApplication(xlsx_data,'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            attachment.add_header('Content-Disposition', 'attachment', filename=f"rapport.xlsx")
            message.attach(attachment)

        try:
            credential = ClientSecretCredential(AD_TENANT_ID, AD_CLIENT_ID, AD_CLIENT_SECRET)
            client = GraphClient(credential=credential)         # Get a authenticated Graph client. Might be better to have one for the whole worker?
            userPrincipalName = "noreply@haninge.se"            # This is the user account that our mail are sent from

            result = client.post(f"/users/{userPrincipalName}/sendMail", 
                                data=base64.b64encode(message.as_string().encode('utf-8')),
                                headers={'Content-Type': 'text/plain'})
            if 'error' in result:
                loggtext = f"sendMail failed! {result['error']['code']: {result['error']['message']}}"
                logging.error(loggtext)
        except Exception as e:         # Some error
            loggtext = f"Send mail failed with error: {e}"    
            logging.error(loggtext)

""" 
Main function
"""
def main():
    es = Elasticsearch(ES_URL)
    if not es.ping():
        logging.fatal(f"Elasticsearch is not reachable at {ES_URL}")
        return

    jinja_env = Environment(loader=FileSystemLoader("."))
    report_template = jinja_env.get_template("report.jinja2.html")

    today = date.today()
    today = date(2022,12,19)
    # today = date(2023,1,1)
    if today.day == 1:      # First day of the month. Run a montly report on the previous mounth
        startday = date(today.year, today.month-1, 1) if today.month != 1 else date(today.year-1, 12, 1)
        datelist = [(startday+timedelta(days=d)).isoformat() for d in range(0,(today-startday).days)]
        data = get_data(datelist)
        header = f"Camunda månadsrapport för {startday.strftime('%B %Y')}"
        html = mk_html(data, header, report_template)
        xlsx = mk_excel(data, startday.strftime('%Y-%m'))
        send_mail(header, html, xlsx)

    if today.weekday() == 0:        # Today is Monday. Run a report on the previous week
        datelist = [(today+timedelta(days=d)).isoformat() for d in range(-7,0)]
        data = get_data(datelist)
        header = f"Camunda veckorapport för vecka {(today+timedelta(days=-1)).strftime('%W %Y')}"
        html = mk_html(data, header, report_template)
        xlsx = mk_excel(data, (today+timedelta(days=-1)).strftime('%W %Y'))
        send_mail(header, html, xlsx)

    yesterday = (today+timedelta(-1)).isoformat()
    data = get_data([yesterday])      # Always run a report on the previous day
    header = f"Camunda dygnsrapport för {yesterday}"
    html = mk_html(data, header, report_template)
    send_mail(header, html)


""" 
Starting point
"""
if __name__ == "__main__":
    if DEBUG_MODE:       # Debug requested
        logging.basicConfig(level=logging.DEBUG, format=LOGFORMAT)
    logging.basicConfig(level=logging.ERROR, format=LOGFORMAT)     # Default logging level

    main()
